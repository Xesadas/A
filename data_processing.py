import os
import logging
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from dash import dcc
import io

# Configuração de logging detalhada
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configuração de caminhos dinâmica
MOUNT_PATH = '/data' if os.environ.get('RENDER') else os.path.join(os.getcwd(), 'data')
EXCEL_PATH = os.path.join(MOUNT_PATH, 'b.xlsx')

def setup_persistent_environment():
    try:
        os.makedirs(MOUNT_PATH, exist_ok=True)

        if not os.path.exists(EXCEL_PATH):
            wb = Workbook()
            
            # Remove sheet padrão vazio se existir
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Cria aba JAN com cabeçalhos
            ws = wb.create_sheet("JAN")
            headers = [
                'data', 'beneficiario', 'valor_transacionado', 'valor_liberado',
                'taxa_de_juros', 'comissao_agente', 'extra_agente', 'valor_dualcred',
                'nota_fiscal', 'porcentagem_agente', 'quantidade_parcelas', 'agente',
                '%trans', '%liberad'
            ]
            ws.append(headers)
            
            # Cria outras abas mensais vazias
            months = ['FEV', 'MAR', 'ABR', 'MAI', 'JUN', 
                     'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
            for month in months:
                ws = wb.create_sheet(month)
                ws.append(headers)
            
            wb.save(EXCEL_PATH)
        
        if not os.access(MOUNT_PATH, os.W_OK):
            logger.error(f"Sem permissão de escrita em: {MOUNT_PATH}")
            raise PermissionError("Erro de permissão no diretório persistente")

    except Exception as e:
        logger.error(f"Falha na configuração inicial: {str(e)}")
        raise

def sanitize_column_name(col):
    return (
        str(col)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("ç", "c")
        .replace("ã", "a")
        .replace("õ", "o")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("à", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("ú", "u")
        .replace("%", "porcento")
        .replace("(", "")
        .replace(")", "")
    )

def load_and_process_data():
    """Carrega dados existentes sem reprocessar cálculos históricos"""
    try:
        logger.info("Carregando dados existentes...")
        setup_persistent_environment()

        # Mapeamento de colunas
        column_mapping = {
            'beneficiário': 'beneficiario',
            'comissão_agente': 'comissao_agente',
            'chave_pix_cpf': 'chave_pix',
            '%_trans': '%trans',
            '%_liberad': '%liberad',
            'máquina': 'maquina'
        }

        # Carregar abas como dicionário de DataFrames
        try:
            sheets = pd.read_excel(EXCEL_PATH, sheet_name=None, engine='openpyxl')
        except FileNotFoundError:
            sheets = {}

        # Apenas sanitizar colunas, sem recálculos
        processed_sheets = {}
        for sheet_name, df in sheets.items():
            try:
                # Sanitizar e padronizar colunas
                df.columns = [sanitize_column_name(col) for col in df.columns]
                df.rename(columns=column_mapping, inplace=True, errors='ignore')
                
                # Garantir colunas necessárias
                required_columns = [
                    'data', 'beneficiario', 'valor_transacionado', 'valor_liberado',
                    'taxa_de_juros', 'comissao_agente', 'extra_agente', 'valor_dualcred',
                    'nota_fiscal', 'porcentagem_agente', 'quantidade_parcelas', 'agente',
                    '%trans', '%liberad'
                ]
                
                for col in required_columns:
                    if col not in df.columns:
                        df[col] = pd.NaT if col == 'data' else 0.0

                processed_sheets[sheet_name] = df
                logger.info(f"Aba {sheet_name} carregada com sucesso")

            except Exception as e:
                logger.error(f"Erro na aba {sheet_name}: {str(e)}")
                continue

        return processed_sheets

    except Exception as e:
        logger.error(f"Falha crítica ao carregar: {str(e)}")
        return {}

def processar_novos_dados(novos_dados):
    """Processa cálculos apenas para novos registros"""
    try:
        logger.info("Processando novos dados...")
        
        # Cálculos para novos registros
        novos_dados['valor_dualcred'] = (
            novos_dados['valor_transacionado'] 
            - novos_dados['valor_liberado'] 
            - novos_dados['taxa_de_juros'] 
            - novos_dados['comissao_agente'] 
            - novos_dados['extra_agente']
        ).round(2)

        novos_dados['%trans'] = np.where(
            novos_dados['valor_transacionado'] > 0,
            (novos_dados['valor_dualcred'] / novos_dados['valor_transacionado']) * 100,
            0
        ).round(2)

        novos_dados['%liberad'] = np.where(
            novos_dados['valor_liberado'] > 0,
            (novos_dados['valor_dualcred'] / novos_dados['valor_liberado']) * 100,
            0
        ).round(2)

        novos_dados['nota_fiscal'] = (novos_dados['valor_transacionado'] * 0.032).round(2)

        return novos_dados

    except Exception as e:
        logger.error(f"Erro no processamento: {str(e)}")
        return pd.DataFrame()

def salvar_no_excel(novos_dados):
    """Adiciona novos dados ao arquivo Excel existente de forma incremental"""
    try:
        logger.info("Iniciando salvamento incremental...")
        setup_persistent_environment()

        # Mapeamento de meses para nomes das abas
        month_names = {
            1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR', 5: 'MAI', 6: 'JUN',
            7: 'JUL', 8: 'AGO', 9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
        }

        # Processar cálculos apenas para novos dados
        novos_dados = processar_novos_dados(novos_dados)
        if novos_dados.empty:
            logger.warning("Nenhum dado novo para salvar")
            return False

        # Converter e mapear datas
        novos_dados['data'] = pd.to_datetime(novos_dados['data'])
        novos_dados['month'] = novos_dados['data'].dt.month.map(month_names)

        # Carregar dados existentes
        try:
            with pd.ExcelFile(EXCEL_PATH, engine='openpyxl') as excel:
                existing_sheets = {sheet: pd.read_excel(excel, sheet_name=sheet) for sheet in excel.sheet_names}
        except FileNotFoundError:
            existing_sheets = {}

        # Atualizar cada aba com novos dados
        with pd.ExcelWriter(
            EXCEL_PATH, 
            engine='openpyxl', 
            mode='a', 
            if_sheet_exists='overlay'
        ) as writer:
            
            for sheet_name, month_num in month_names.items():
                sheet_name = month_names[month_num]
                
                # Filtrar novos dados para a aba
                novos_para_aba = novos_dados[novos_dados['month'] == sheet_name].drop(columns=['month'])
                
                if not novos_para_aba.empty:
                    # Carregar dados existentes
                    existing_df = existing_sheets.get(sheet_name, pd.DataFrame())
                    
                    # Combinar dados
                    combined_df = pd.concat([existing_df, novos_para_aba], ignore_index=True)
                    
                    # Remover duplicatas
                    combined_df = combined_df.drop_duplicates(subset=['data', 'beneficiario', 'valor_transacionado'])
                    
                    # Ordenar por data
                    combined_df = combined_df.sort_values(by='data').reset_index(drop=True)
                    
                    # Salvar aba atualizada
                    combined_df.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False,
                        columns=[
                            'data', 'beneficiario', 'valor_transacionado', 'valor_liberado',
                            'taxa_de_juros', 'comissao_agente', 'extra_agente', 'valor_dualcred',
                            'nota_fiscal', 'porcentagem_agente', 'quantidade_parcelas', 'agente',
                            '%trans', '%liberad'
                        ]
                    )

        logger.info("Dados salvos com sucesso (incremental)")
        return True

    except Exception as e:
        logger.error(f"Falha ao salvar incrementalmente: {str(e)}", exc_info=True)
        return False

def exportar_dados(processed_sheets):
    """Exporta mantendo a estrutura por abas"""
    try:
        logger.info("Iniciando exportação...")
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            for sheet_name, df in processed_sheets.items():
                logger.info(f"Exportando aba: {sheet_name}")
                
                if df.empty:
                    logger.warning(f"Aba {sheet_name} vazia")
                    continue
                    
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    columns=[
                        'data', 'beneficiario', 'valor_transacionado', 'valor_liberado',
                        'taxa_de_juros', 'comissao_agente', 'extra_agente', 'valor_dualcred',
                        'nota_fiscal', 'quantidade_parcelas', 'agente', '%trans', '%liberad'
                    ]
                )
        
        buffer.seek(0)
        logger.info("Exportação concluída com sucesso")
        return dcc.send_bytes(
            buffer.getvalue(),
            filename="Dados_Atualizados.xlsx",
            type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except Exception as e:
        logger.error(f"Erro na exportação: {str(e)}", exc_info=True)
        return None

# Inicialização segura
try:
    processed_sheets = load_and_process_data()
    if not processed_sheets:
        logger.warning("Nenhuma aba válida encontrada")
    else:
        logger.info(f"Dados carregados: {len(processed_sheets)} abas")
except Exception as e:
    logger.error(f"Falha crítica: {str(e)}")
    processed_sheets = {}